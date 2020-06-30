from telegram import InlineKeyboardButton, InlineKeyboardMarkup, ParseMode
from telegram.error import BadRequest

from openpyxl import Workbook


import math
import os
from datetime import datetime, timedelta
from crawler import get_matched_posts_database, get_last_post_url

from bot import SELECTING_ACTION, start_markup, home

from persiantools.jdatetime import JalaliDate

from login import login_required

from models import Channel, Admin, post_format

SET_QUERY, SET_DATE, GET_POSTS = range(7, 10)


# @login_required
def query_alert(update, context):
    update.message.reply_text('متن جستجو:')
    return SET_QUERY


# @login_required
def start_process(update, context):
    context.user_data['me'] = Admin.get_by_username(
        update.message.from_user.username)
    context.user_data['query'] = update.effective_message.text
    if Channel.get_all() == []:
        update.message.reply_text(
            'شما ابتدا باید برای جستجو کانالی تعیین کنید')
        return SELECTING_ACTION

    keyboard = [[InlineKeyboardButton("امروز", callback_data='1'),
                 InlineKeyboardButton("یک هفته اخیر", callback_data='2')],
                [InlineKeyboardButton("یک ماه اخیر", callback_data='3')],
                [InlineKeyboardButton("سه ماه اخیر", callback_data='4')],
                [InlineKeyboardButton("بازگشت به خانه", callback_data='0')]
                ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    update.message.reply_text(
        'بازه‌ی جستجو را مشخص کنید', reply_markup=reply_markup)

    return SET_DATE


# def manually_date_alert(update, context):
#     update.callback_query.answer()
#     update.callback_query.edit_message_text(
#         " تاریخ را با فرمت زیر ارسال کنید \n\n"
#         "روز-ماه-سال\n"
#         "مثال: 05-02-1399")
#     return SET_DATE


# def manually_date(update, context):
#     persian_date = list(map(int, update.effective_message.text.split('-')))
#     try:
#         date = JalaliDate(persian_date[0], persian_date[1],
#                           persian_date[2]).to_gregorian()
#     except ValueError:
#         update.message.reply_text('تاریخ را درست وارد کنید')
#         return SET_DATE
#     update.message.reply_text('لطفا صبر کنید...')

#     return get_posts(update, context, date)


# def wrong_date(update, context):
#     update.message.reply_text('تاریخ را درست وارد کنید')
#     return SET_DATE


def choose_date(update, context):
    update.callback_query.answer()
    data = update.callback_query.data
    if data == '1':
        date = datetime.today().date()
    elif data == '2':
        date = (datetime.today() - timedelta(days=7)).date()
    elif data == '3':
        date = (datetime.today() - timedelta(days=30)).date()
    elif data == '4':
        date = (datetime.today() - timedelta(days=90)).date()
    update.callback_query.edit_message_text('لطفا صبر کنید...')

    return get_posts(update.callback_query, context, date)


def get_posts(update, context, date):
    context.user_data['all_posts'] = []

    for post in get_matched_posts_database(context.user_data['query'], date):
        context.user_data['all_posts'].append(post)
    context.user_data['posts_count'] = len(context.user_data['all_posts'])
    make_excel_file(posts=context.user_data['all_posts'],
                    query=context.user_data['query'])
    return next_posts(update, context)


def next_posts(update, context):
    keyboard = [[InlineKeyboardButton("ریست و بازگشت به خانه", callback_data='5')],
                [InlineKeyboardButton("صفحه بعد", callback_data='1')],
                [InlineKeyboardButton("دریافت فایل اکسل", callback_data='6')]]
    last_keyboard = [
        [InlineKeyboardButton("ریست و بازگشت به خانه", callback_data='5')],
        [InlineKeyboardButton("دریافت فایل اکسل", callback_data='6')]
    ]
    if not update.message:
        update = update.callback_query

    count = len(context.user_data['all_posts'])
    pages = math.ceil(context.user_data['posts_count'] / 5)

    if count == 0:
        update.message.reply_text(
            'متأسفانه پستی یافت نشد', reply_markup=start_markup)
        return home(update, context)
    elif count <= 5:
        show_count = count
        show_posts(show_count, InlineKeyboardMarkup(
            last_keyboard), pages, count, update, context)
        return GET_POSTS
    elif count > 5:
        show_posts(5, InlineKeyboardMarkup(keyboard),
                   pages, count, update, context)
        return GET_POSTS


def show_posts(show_count, reply_markup, pages, count, update, context):
    if count < 5:
        current_page = pages - int(count/5)
        last = context.user_data['posts_count']
    elif count % 5 == 0:
        current_page = pages - int(count/5) + 1
        last = current_page*5
    else:
        current_page = pages - int(count/5)
        last = current_page*5
    for post in context.user_data['all_posts'][0:show_count]:
        try:
            update.message.reply_text(
                prettify(post), parse_mode=ParseMode.MARKDOWN)
        except BadRequest:
            continue
    context.user_data['all_posts'] = context.user_data['all_posts'][show_count:]
    update.message.reply_text("{0} پست یافت شد \n\n"
                              "صفحه {1} از {2} \n\n"
                              "شماره های {3} تا {4}".format(
                                  context.user_data['posts_count'], current_page, pages, (current_page-1)*5 + 1, last), reply_markup=reply_markup)


def prettify(post):
    prettier = "📢 {0} \n\n"\
        "🔗 [مشاهده پست در کانال]({1}) \n\n"\
        "👁‍🗨 تعداد بازدید: {2} \n➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖➖\n"\
        "💬 کپشن: \n {3} \n\n".format(
            post['channel_name'], post['url'], post['views'], post['caption'])
    return prettier


def make_excel_file(posts, query):
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    i = 2

    ws.cell(1, 1).value = 'شماره تولید'
    ws.cell(1, 2).value = 'نام برنامه'
    ws.cell(1, 3).value = 'عنوان'
    ws.cell(1, 4).value = 'نوع'
    ws.cell(1, 5).value = 'مدت(ثانیه)'
    ws.cell(1, 6).value = 'نام کانال/صفحه'
    ws.cell(1, 7).value = 'لینک 24 ساعته'
    ws.cell(1, 8).value = 'تعداد بازدید'
    ws.cell(1, 9).value = 'زمان انتشار'

    for post in posts:
        ws.cell(i, 3).value = post['caption']

        ws.cell(i, 4).value = post_format[post['format']]

        ws.cell(i, 5).value = post['duration']

        ws.cell(i, 6).value = post['channel_name']

        ws.cell(i, 7).value = post['url']
        ws.cell(i, 7).hyperlink = post['url']

        if post['views'] == 0:
            ws.cell(i, 8).value = 'تعداد ویوها مشخص نشد'
        else:
            ws.cell(i, 8).value = post['views']
        ws.cell(i, 8).number_format = '0'

        ws.cell(i, 9).value = post['datetime']

        i += 1

    filename = '{}.xlsx'.format(query)
    wb.save(filename=filename)

    return filename


def send_excel(update, context):
    print('fetch')
    update = update.callback_query
    update.answer()
    temp = update.message.reply_text('در حال بارگذاری ...')
    update.message.reply_document(document=open(
        '{}.xlsx'.format(context.user_data['query']), 'rb'))
    update.message.bot.delete_message(
        chat_id=update.message.chat_id, message_id=temp.message_id)

    return home(update, context)
